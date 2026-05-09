import cv2
import numpy as np
import pytesseract
from PIL import Image
import pandas as pd
import re
from pathlib import Path
from typing import Tuple, List
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# --- CONFIG ---
IMAGE_DIR = r"C:\Users\Megaport\canvas_images"
IMAGE_BASENAME = "buy_1254535974526922850"  # sans extension
OUTPUT_XLSX = "output.xlsx"
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


# --- UTILITAIRES ---
def find_image_file(directory: str, basename: str) -> str:
    p = Path(directory)
    if not p.exists() or not p.is_dir():
        raise FileNotFoundError(f"Répertoire introuvable : {directory}")
    exts = [".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp"]
    candidate = p / basename
    if candidate.exists():
        return str(candidate)
    for e in exts:
        f = p / (basename + e)
        if f.exists():
            return str(f)
    for f in p.iterdir():
        if f.is_file() and f.stem == basename:
            return str(f)
    raise FileNotFoundError(f"Aucun fichier trouvé pour '{basename}' dans {directory}")


def read_image_or_raise(path: str) -> np.ndarray:
    img = cv2.imread(path)
    if img is None:
        raise FileNotFoundError(f"Impossible de lire l'image : {path}")
    return img


def preprocess(img_path: str) -> Tuple[np.ndarray, np.ndarray]:
    img = read_image_or_raise(img_path)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    _, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if np.mean(th) < 127:
        th = 255 - th
    return img, th


def detect_table_cells(bin_img: np.ndarray) -> List[Tuple[int, int, int, int]]:
    kernel_len = max(
        3, bin_img.shape[1] // 300
    )  # Ajustement pour une détection plus fine
    ver_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_len))
    hor_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_len, 1))
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))

    vertical = cv2.erode(bin_img, ver_kernel, iterations=3)
    vertical = cv2.dilate(vertical, ver_kernel, iterations=3)

    horizontal = cv2.erode(bin_img, hor_kernel, iterations=3)
    horizontal = cv2.dilate(horizontal, hor_kernel, iterations=3)

    table_mask = cv2.addWeighted(vertical, 0.5, horizontal, 0.5, 0.0)
    table_mask = cv2.erode(cv2.bitwise_not(table_mask), kernel, iterations=2)
    _, table_mask = cv2.threshold(
        table_mask, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU
    )

    contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    boxes: List[Tuple[int, int, int, int]] = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        if w > 15 and h > 15:  # Filtrer les petites boîtes
            boxes.append((int(x), int(y), int(w), int(h)))

    boxes = [b for b in boxes if b[2] > 10 and b[3] > 10]
    boxes = sorted(boxes, key=lambda x: (x[1], x[0]))
    return boxes


def group_boxes_to_rows(
    boxes: List[Tuple[int, int, int, int]], tol: int = 10
) -> List[List[Tuple[int, int, int, int]]]:
    rows = []
    current_row = []
    last_y = -999
    for b in boxes:
        x, y, w, h = b
        if last_y == -999 or abs(y - last_y) <= tol:
            current_row.append(b)
            last_y = y if last_y == -999 else (last_y + y) // 2
        else:
            rows.append(sorted(current_row, key=lambda r: r[0]))
            current_row = [b]
            last_y = y
    if current_row:
        rows.append(sorted(current_row, key=lambda r: r[0]))
    return rows


def ocr_cell(img: np.ndarray, box: Tuple[int, int, int, int]) -> str:
    x, y, w, h = box
    pad = 2
    crop = img[max(0, y - pad) : y + h + pad, max(0, x - pad) : x + w + pad]
    pil = Image.fromarray(cv2.cvtColor(crop, cv2.COLOR_BGR2RGB))
    txt = pytesseract.image_to_string(pil, config="--psm 6")
    txt = txt.strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt


# --- Normalisation et appariement prix/quantité ---
def pair_prices_quantities_from_two_rows(
    header_row: List[str], value_row: List[str]
) -> List[tuple]:
    if header_row and header_row[0].lower().startswith("prix"):
        prices = header_row[1:]
    else:
        prices = header_row[:]

    if value_row and value_row[0].lower().startswith("quant"):
        quantities = value_row[1:]
    else:
        quantities = value_row[:]

    def clean_price(p: str) -> str:
        p = p.strip()
        p = re.sub(r"(\d)\.(\d)", r"\1,\2", p)
        return p

    def clean_qty(q: str) -> str:
        q = q.strip()
        q = re.sub(r"[^\d\-]", "", q)
        return q

    prices = [clean_price(p) for p in prices if str(p).strip()]
    quantities = [clean_qty(q) for q in quantities if str(q).strip()]

    pairs = list(zip(prices, quantities))
    if len(prices) != len(quantities):
        print(
            f"Avertissement: nombre prix ({len(prices)}) != nombre quantités ({len(quantities)}). Troncature appliquée."
        )
    return pairs


def normalize_to_two_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Cas où tout est dans une seule cellule (ex: "Prix $0,12 $0,11 ... Quant 23 35 ...")
    if df.shape[0] == 1 and df.shape[1] == 1:
        text = df.iloc[0, 0]
        if "Prix" in text and ("Quant" in text or "quantité" in text):
            # Extraire les prix
            prix_part = text[text.find("Prix") : text.find("Quant")].strip()
            prices = re.findall(r"\$\d+[\.,]\d+", prix_part)
            prices = [re.sub(r"(\d)\.(\d)", r"\1,\2", p) for p in prices]

            # Extraire les quantités
            quant_part = text[text.find("Quant") :].strip()
            quantities = re.findall(r"\d+", quant_part)

            if len(prices) == len(quantities):
                return pd.DataFrame({"Prix buy": prices, "Quantités buy": quantities})

    # Cas où il y a deux lignes (prix et quantités)
    if df.shape[0] == 2 and df.shape[1] == 1:
        header_text = df.iloc[0, 0]
        value_text = df.iloc[1, 0]
        if "Prix" in header_text and (
            "Quant" in value_text or "quantité" in value_text
        ):
            prices = re.findall(r"\$\d+[\.,]\d+", header_text)
            prices = [re.sub(r"(\d)\.(\d)", r"\1,\2", p) for p in prices]
            quantities = re.findall(r"\d+", value_text)
            if len(prices) == len(quantities):
                return pd.DataFrame({"Prix buy": prices, "Quantités buy": quantities})

    # Cas où il y a deux lignes et plusieurs colonnes
    if df.shape[0] == 2 and df.shape[1] >= 2:
        header_row = df.iloc[0].astype(str).tolist()
        value_row = df.iloc[1].astype(str).tolist()
        pairs = pair_prices_quantities_from_two_rows(header_row, value_row)
        if pairs:
            prices, quantities = zip(*pairs)
            return pd.DataFrame(
                {"Prix buy": list(prices), "Quantités buy": list(quantities)}
            )

    # Cas où les données sont déjà en deux colonnes
    if df.shape[1] == 2:
        df2 = df.copy()
        df2.columns = ["Prix buy", "Quantités buy"]
        df2["Prix buy"] = (
            df2["Prix buy"]
            .astype(str)
            .apply(lambda p: re.sub(r"(\d)\.(\d)", r"\1,\2", p.strip()))
        )
        df2["Quantités buy"] = (
            df2["Quantités buy"]
            .astype(str)
            .str.replace(r"[^\d\-]", "", regex=True)
            .str.strip()
        )
        return df2

    # Cas par défaut : essayer de trouver des paires prix/quantité
    flat = [str(x).strip() for x in df.values.flatten() if str(x).strip()]
    lower_flat = [s.lower() for s in flat]
    if "prix" in lower_flat and ("quantite" in lower_flat or "quantité" in lower_flat):
        try:
            i_prix = lower_flat.index("prix")
            i_quant = next(i for i, s in enumerate(lower_flat) if s.startswith("quant"))
            prices = flat[i_prix + 1 : i_quant]
            quantities = flat[i_quant + 1 :]
            pairs = list(
                zip(
                    [re.sub(r"(\d)\.(\d)", r"\1,\2", p) for p in prices],
                    [re.sub(r"[^\d\-]", "", q) for q in quantities],
                )
            )
            if pairs:
                prices, quantities = zip(*pairs)
                return pd.DataFrame(
                    {"Prix buy": list(prices), "Quantités buy": list(quantities)}
                )
        except StopIteration:
            pass

    prices = []
    quantities = []
    i = 0
    while i < len(flat) - 1:
        a = flat[i]
        b = flat[i + 1]
        if re.search(r"^\$?\d+[.,]\d+", a) and re.search(r"\d+", b):
            prices.append(re.sub(r"(\d)\.(\d)", r"\1,\2", a))
            quantities.append(re.sub(r"[^\d\-]", "", b))
            i += 2
        else:
            i += 1
    if prices:
        return pd.DataFrame({"Prix buy": prices, "Quantités buy": quantities})

    df.columns = ["Prix buy"] + [f"col{i}" for i in range(2, df.shape[1] + 1)]
    return df


# --- STYLE EXCEL ---
def style_excel_table(path: str, sheet_name: str = "Sheet1"):
    wb = load_workbook(path)
    ws_candidate = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if isinstance(ws_candidate, Worksheet):
        ws = ws_candidate
    else:
        ws = wb.worksheets[0]

    header_font = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font

    col_widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            col_widths[col] = max(col_widths.get(col, 0), len(str(cell.value)) + 2)
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 2:
        last_col_letter = get_column_letter(max_col)
        table_ref = f"A1:{last_col_letter}{max_row}"
        tab = Table(displayName="Table1", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        existing_names = (
            [t.displayName for t in ws._tables] if hasattr(ws, "_tables") else []
        )
        if "Table1" not in existing_names:
            ws.add_table(tab)

    wb.save(path)


# --- MAIN ---
def extract_table_to_excel(image_path: str, output_xlsx: str) -> None:
    orig, bin_img = preprocess(image_path)
    boxes = detect_table_cells(bin_img)
    if not boxes:
        text = pytesseract.image_to_string(Image.open(image_path), config="--psm 6")
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        rows = [re.split(r"\s{2,}", l) for l in lines]
        if not rows:
            raise RuntimeError(
                "Aucune table détectée et OCR fallback n'a pas trouvé de lignes valides."
            )
        df = pd.DataFrame(rows)
        final = normalize_to_two_columns(df)
        final.to_excel(output_xlsx, index=False)
        style_excel_table(output_xlsx)
        return

    rows_boxes = group_boxes_to_rows(boxes, tol=12)
    table = []
    for row in rows_boxes:
        row_texts = [ocr_cell(orig, b) for b in row]
        table.append(row_texts)

    max_cols = max(len(r) for r in table)
    norm_table = [r + [""] * (max_cols - len(r)) for r in table]
    df = pd.DataFrame(norm_table)

    final_df = normalize_to_two_columns(df)
    final_df["Quantités buy"] = final_df["Quantités buy"].replace("", pd.NA)
    try:
        final_df["Quantités buy"] = final_df["Quantités buy"].astype("Int64")
    except Exception:
        pass

    final_df.to_excel(output_xlsx, index=False)
    style_excel_table(output_xlsx)


if __name__ == "__main__":
    try:
        image_path = find_image_file(IMAGE_DIR, IMAGE_BASENAME)
        print(f"Image trouvée : {image_path}")
        extract_table_to_excel(image_path, OUTPUT_XLSX)
        print("Export terminé :", OUTPUT_XLSX)
    except FileNotFoundError as e:
        print("Erreur :", e)
    except Exception as e:
        print("Erreur inattendue :", str(e))

    # end main
